import os
import json
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def collect_instruction_json_stats(root_dir):
    error_counter = Counter()
    warning_counter = Counter()
    type_counter = Counter()

    file_summaries = []
    issue_details = []
    invalid_files = []

    for current_root, _, files in os.walk(root_dir):
        for file_name in files:
            if file_name.startswith("instruction") and file_name.endswith(".json"):
                file_path = os.path.join(current_root, file_name)

                try:
                    with open(file_path, "r", encoding="utf-8") as f:
                        data = json.load(f)

                    project = data.get("project", os.path.splitext(file_name)[0])
                    issues = data.get("issues", [])

                    file_error_counter = Counter()
                    file_warning_counter = Counter()
                    file_type_counter = Counter()

                    for issue in issues:
                        issue_type = str(issue.get("type", "unknown")).lower()
                        symbol = issue.get("symbol", "unknown")

                        type_counter[issue_type] += 1
                        file_type_counter[issue_type] += 1

                        if issue_type == "error":
                            error_counter[symbol] += 1
                            file_error_counter[symbol] += 1
                        elif issue_type == "warning":
                            warning_counter[symbol] += 1
                            file_warning_counter[symbol] += 1

                        issue_details.append({
                            "project": project,
                            "json_file": file_name,
                            "json_path": file_path,
                            "issue_type": issue_type,
                            "symbol": symbol,
                            "message": issue.get("message", ""),
                            "message_id": issue.get("message-id", ""),
                            "module": issue.get("module", ""),
                            "obj": issue.get("obj", ""),
                            "line": issue.get("line", ""),
                            "column": issue.get("column", ""),
                            "code_path": issue.get("path", "")
                        })

                    file_summaries.append({
                        "project": project,
                        "json_file": file_name,
                        "json_path": file_path,
                        "issue_total": len(issues),
                        "error_count": sum(file_error_counter.values()),
                        "warning_count": sum(file_warning_counter.values()),
                        "convention_count": file_type_counter.get("convention", 0),
                        "refactor_count": file_type_counter.get("refactor", 0),
                        "info_count": file_type_counter.get("info", 0),
                    })

                except Exception as e:
                    invalid_files.append({
                        "json_file": file_name,
                        "json_path": file_path,
                        "error": str(e)
                    })

    result = {
        "total_files": len(file_summaries),
        "total_error_count": sum(error_counter.values()),
        "total_warning_count": sum(warning_counter.values()),
        "error_by_symbol": dict(sorted(error_counter.items(), key=lambda x: (-x[1], x[0]))),
        "warning_by_symbol": dict(sorted(warning_counter.items(), key=lambda x: (-x[1], x[0]))),
        "all_issue_type_count": dict(sorted(type_counter.items(), key=lambda x: (-x[1], x[0]))),
        "file_summaries": file_summaries,
        "issue_details": issue_details,
        "invalid_files": invalid_files,
    }
    return result


def style_worksheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border


    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    ws.freeze_panes = "A2"


def export_to_excel(result, output_excel):
    wb = Workbook()


    default_ws = wb.active
    wb.remove(default_ws)


    ws = wb.create_sheet("summary")
    ws.append(["metic", "value"])
    ws.append([" JSON file num", result["total_files"]])
    ws.append(["Error num", result["total_error_count"]])
    ws.append(["Warning num", result["total_warning_count"]])

    for issue_type, count in result["all_issue_type_count"].items():
        ws.append([f"{issue_type} num", count])

    if result["invalid_files"]:
        ws.append(["failed file num", len(result["invalid_files"])])
    else:
        ws.append(["failed file num ", 0])

    style_worksheet(ws)


    ws = wb.create_sheet("error_breakdown")
    ws.append(["error_symbol", "count"])
    for symbol, count in result["error_by_symbol"].items():
        ws.append([symbol, count])
    style_worksheet(ws)

  
    ws = wb.create_sheet("warning_breakdown")
    ws.append(["warning_symbol", "count"])
    for symbol, count in result["warning_by_symbol"].items():
        ws.append([symbol, count])
    style_worksheet(ws)

    ws = wb.create_sheet("file_summary")
    ws.append([
        "project", "json_file", "json_path",
        "issue_total", "error_count", "warning_count",
        "convention_count", "refactor_count", "info_count"
    ])
    for item in result["file_summaries"]:
        ws.append([
            item["project"],
            item["json_file"],
            item["json_path"],
            item["issue_total"],
            item["error_count"],
            item["warning_count"],
            item["convention_count"],
            item["refactor_count"],
            item["info_count"],
        ])
    style_worksheet(ws)


    ws = wb.create_sheet("issue_details")
    ws.append([
        "project", "json_file", "json_path", "issue_type", "symbol",
        "message", "message_id", "module", "obj", "line", "column", "code_path"
    ])
    for item in result["issue_details"]:
        ws.append([
            item["project"],
            item["json_file"],
            item["json_path"],
            item["issue_type"],
            item["symbol"],
            item["message"],
            item["message_id"],
            item["module"],
            item["obj"],
            item["line"],
            item["column"],
            item["code_path"],
        ])
    style_worksheet(ws)


    if result["invalid_files"]:
        ws = wb.create_sheet("invalid_files")
        ws.append(["json_file", "json_path", "error"])
        for item in result["invalid_files"]:
            ws.append([item["json_file"], item["json_path"], item["error"]])
        style_worksheet(ws)

    wb.save(output_excel)


if __name__ == "__main__":

    # files have been analysised path 
    root_dir = ""
    # report path
    output_excel = "lint/deepseek/reports.xlsx"

    result = collect_instruction_json_stats(root_dir)
    export_to_excel(result, output_excel)
